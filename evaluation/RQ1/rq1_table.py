#!/usr/bin/env python3
"""RQ1 results table — built solely from rq1_final.jsonl (single source of truth).

Rows = path-chain class (mutually exclusive). Sum of File column = corpus.
Columns: File / stream / rops file·stream / ps file·stream / ms file·stream

Reachability definition
  file   = did the tool reach **all** oracle streams of that file (completeness)
  stream = fraction of oracle streams reached (sum/sum)
  RoPS   = rops_stream_cover (count cover; opaque preserved as a PGW blob)
  ps/ms  = ps_reached / ms_reached (trace, against oracle sha)
  RoPS-unmeasured files (e.g. A5 probe) are excluded from the RoPS denominator and shown in a footnote.
Usage: python3 rq1_table.py --final rq1_final.jsonl --out RQ1_Results.xlsx
"""
import argparse, json, collections
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AR = 'Arial'
ORD = ['Serialization', 'Opaque (PGW)',
       'Container', 'Container→opaque', 'Container→Streaming', 'Container→Streaming→opaque',
       'Container→Container', 'Container→Container→opaque',
       'Streaming', 'Streaming→opaque', 'Streaming→Streaming', 'Streaming→Streaming→opaque',
       'Mixed', 'No stream']
# Display-only labels for chain_label values kept as data/logic keys.
DISPLAY = {'Mixed': 'Mixed', 'No stream': 'No stream'}


def lsort(k):
    return (ORD.index(k) if k in ORD else 50, k)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--final', required=True)
    ap.add_argument('--out', required=True)
    a = ap.parse_args()
    recs = [json.loads(l) for l in open(a.final, encoding='utf-8') if l.strip()]

    G = collections.defaultdict(list)
    for r in recs:
        G[r['chain_label']].append(r)

    wb = Workbook()
    thin = Side('thin', color='BFBFBF'); bd = Border(thin, thin, thin, thin)
    HDR = PatternFill('solid', fgColor='DDEBF7'); SUB = PatternFill('solid', fgColor='F2F2F2')
    NS = PatternFill('solid', fgColor='FCE4D6')
    ws = wb.active; ws.title = 'Results'
    ws['A1'] = 'RQ1 loading-path reachability — mutually exclusive classes by path chain'
    ws['A1'].font = Font(AR, bold=True, size=13)
    n_ns = sum(1 for r in recs if r['chain_label'] == 'No stream')
    n_unmeas = sum(1 for r in recs if not r['rops_measured'])
    ws['A2'] = (f'Each file attributed to one path chain (mutually exclusive) → sum of File column = corpus {len(recs)}. '
                f'Reachability targets {len(recs)-n_ns} + No stream {n_ns}. '
                f'Terminals (Serialization/Opaque(PGW)) are split by RoPS RFP. '
                + (f'{n_unmeas} RoPS-unmeasured records excluded from the RoPS denominator (footnote).' if n_unmeas else ''))
    ws['A2'].font = Font(AR, italic=True, size=9)

    r0 = 4
    hdr = ['Path-chain class', 'File', 'stream', 'rops file', 'rops stream',
           'ps file', 'ps stream', 'ms file', 'ms stream']
    for j, c in enumerate(hdr, 1):
        x = ws.cell(r0, j, c); x.font = Font(AR, bold=True); x.fill = HDR; x.border = bd
        x.alignment = Alignment('center', 'center', wrap_text=True)

    def frac(num, den):
        return round(num / den, 3) if den else None

    r = r0 + 1
    TOT = dict(F=0, st=0, rf_n=0, rf_d=0, rs_n=0, rs_d=0, pf=0, ps_n=0, mf=0, ms_n=0, meas=0)
    for k in sorted(G, key=lsort):
        g = G[k]; is_ns = (k == 'No stream')
        nf = len(g); nst = sum(r_['n_oracle'] for r_ in g)
        # RoPS (measured files only)
        gm = [r_ for r_ in g if r_['rops_measured']]
        rf_file = frac(sum(1 for r_ in gm if r_['rops_stream_cover'] == 1.0), len(gm)) if gm else None
        rs_num = sum(r_['n_oracle'] for r_ in gm if r_['rops_stream_cover'] == 1.0) \
            + sum(round((r_['rops_stream_cover'] or 0) * r_['n_oracle']) for r_ in gm if r_['rops_stream_cover'] != 1.0)
        rs_den = sum(r_['n_oracle'] for r_ in gm)
        rops_stream = frac(rs_num, rs_den)
        # ps / ms (completeness: all streams reached)
        pf = frac(sum(1 for r_ in g if r_['n_oracle'] and r_['ps_reached'] == r_['n_oracle']),
                  sum(1 for r_ in g if r_['n_oracle']))
        mf = frac(sum(1 for r_ in g if r_['n_oracle'] and r_['ms_reached'] == r_['n_oracle']),
                  sum(1 for r_ in g if r_['n_oracle']))
        ps_stream = frac(sum(r_['ps_reached'] for r_ in g), nst)
        ms_stream = frac(sum(r_['ms_reached'] for r_ in g), nst)

        klabel = DISPLAY.get(k, k)
        if is_ns:
            vals = [klabel, nf, 0, '—', '—', '—', '—', '—', '—']
        else:
            vals = [klabel, nf, nst,
                    rf_file if rf_file is not None else '—', rops_stream if rops_stream is not None else '—',
                    pf if pf is not None else '—', ps_stream if ps_stream is not None else '—',
                    mf if mf is not None else '—', ms_stream if ms_stream is not None else '—']
        for j, v in enumerate(vals, 1):
            x = ws.cell(r, j, v); x.border = bd; x.font = Font(AR)
            if j >= 2:
                x.alignment = Alignment('center')
            if j >= 4 and isinstance(v, float):
                x.number_format = '0.000'
            if is_ns:
                x.fill = NS
        if not is_ns:
            TOT['F'] += nf; TOT['st'] += nst
            TOT['rs_n'] += rs_num; TOT['rs_d'] += rs_den; TOT['meas'] += len(gm)
            TOT['rf_n'] += sum(1 for r_ in gm if r_['rops_stream_cover'] == 1.0)
            TOT['ps_n'] += sum(r_['ps_reached'] for r_ in g); TOT['ms_n'] += sum(r_['ms_reached'] for r_ in g)
            TOT['pf'] += sum(1 for r_ in g if r_['n_oracle'] and r_['ps_reached'] == r_['n_oracle'])
            TOT['mf'] += sum(1 for r_ in g if r_['n_oracle'] and r_['ms_reached'] == r_['n_oracle'])
        else:
            TOT['F'] += nf
        r += 1

    n_reach_files = sum(1 for r_ in recs if r_['n_oracle'])
    tv = ['Total', TOT['F'], TOT['st'],
          frac(TOT['rf_n'], TOT['meas']), frac(TOT['rs_n'], TOT['rs_d']),
          frac(TOT['pf'], n_reach_files), frac(TOT['ps_n'], TOT['st']),
          frac(TOT['mf'], n_reach_files), frac(TOT['ms_n'], TOT['st'])]
    for j, v in enumerate(tv, 1):
        x = ws.cell(r, j, v); x.border = bd; x.font = Font(AR, bold=True); x.fill = SUB
        if j >= 2:
            x.alignment = Alignment('center')
        if j >= 4 and isinstance(v, float):
            x.number_format = '0.000'
    r += 2
    ws.cell(r, 1, f'Sum of File column = {TOT["F"]} (equals corpus, mutually exclusive). '
                  f'RoPS stream = count cover (n_rops ≥ oracle). '
                  f'The 8 Opaque(PGW) streams have a sha differing from the harness raw sha and are preserved as PGW blobs, but their content is reached.'
            ).font = Font(AR, italic=True, size=9)
    r += 1
    if n_unmeas:
        um = [r_['sample_id'] for r_ in recs if not r_['rops_measured']]
        ws.cell(r, 1, f'{n_unmeas} RoPS-unmeasured records (excluded from denominator): ' + ' · '.join(um)
                      + ' — craft A5 loading-path probe. Fill via the README top-up command.'
                ).font = Font(AR, italic=True, size=9)
        r += 1
    ns_files = [r_['sample_id'] for r_ in recs if r_['chain_label'] == 'No stream']
    if ns_files:
        ws.cell(r, 1, f'No stream {len(ns_files)} records: '
                      + ' · '.join(f"{s}" for s in ns_files)).font = Font(AR, italic=True, size=9)
    for j, w in enumerate([26, 7, 9, 10, 11, 9, 10, 9, 10], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Analysis sheet
    ws2 = wb.create_sheet('Analysis')
    cols = ['file name', 'tier', 'origin', 'role', 'path chain', 'oracle', 'n_rops',
            'rops reach', 'ps reach', 'ms reach', 'chain raw']
    for j, c in enumerate(cols, 1):
        x = ws2.cell(1, j, c); x.font = Font(AR, bold=True); x.fill = HDR; x.border = bd
        x.alignment = Alignment('center', 'center')
    for i, r_ in enumerate(sorted(recs, key=lambda d: (lsort(d['chain_label']), d['sample_id'])), 2):
        raw = ';'.join(s['chain'] for s in r_['streams']) or '—'
        row = [r_['sample_id'], r_['tier'], r_['origin'], r_['rq1_role'],
               DISPLAY.get(r_['chain_label'], r_['chain_label']),
               r_['n_oracle'], (r_['n_rops'] if r_['rops_measured'] else 'not measured'),
               (r_['rops_stream_cover'] if r_['rops_measured'] else '—'),
               r_['ps_reached'], r_['ms_reached'], raw]
        for j, v in enumerate(row, 1):
            x = ws2.cell(i, j, v); x.font = Font(AR); x.border = bd
            if r_['chain_label'] == 'No stream':
                x.fill = NS
    for j, w in enumerate([16, 8, 20, 12, 22, 8, 8, 10, 8, 8, 18], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = 'A2'
    wb.save(a.out)
    print('->', a.out, '| rows:', {DISPLAY.get(k, k): len(G[k]) for k in sorted(G, key=lsort)})


if __name__ == '__main__':
    main()
