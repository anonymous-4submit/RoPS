#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RQ3 results table — reads only rq3_final.jsonl to build confusion matrices / metrics.

Main table = Real-world 304 (RQ3 comparison target)
Auxiliary table 1 = gadget variants 115 (separate)  ·  Auxiliary table 2 = Combined 419 (merged)

Each block: rows=tools, cols = Judged T · Judged F · Inc · TP · FN · FP · TN · Recall · FNR · FPR · Precision · F1 · Accuracy
   Judged T = number of files the tool judged as T(malicious) (=TP+FP)
   Judged F = number of files the tool judged as F(benign) (=TN + GT-malicious seen as F)
   Inc      = files where the tool could not produce T/F, inconclusive (n/a). 'Inc mal N / ben M' = of those, N malicious · M benign
Confusion matrix: TP=GT-malicious & Judged T  FN=GT-malicious & Judged != T (incl. Inc)  FP=GT-benign & Judged T  TN=GT-benign & Judged F
   inconclusive (n/a): if malicious counted into FN, if benign excluded from the denominator (prevents free TN).
The 'Metric definitions' sheet also records the formulas, judgment criteria, and Inc definition.
Usage: python3 rq3_table.py --final rq3_final.jsonl --out RQ3_results.xlsx
"""
import argparse, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AR = 'Arial'
NA = 'n/a'
TOOLS = [('RoPS (warn-or-above)', 'RoPS', True),
         ('picklescan 1.0.4', 'picklescan', False),
         ('modelscan 0.8.6', 'modelscan', False),
         ('fickling ≥ SUSPICIOUS', 'fickling_loose', False),
         ('weights-only (global-block)', 'weights_only', False)]
ALT = [('RoPS unsafe-only (unsafe)', 'RoPS_confirm', False),
       ('fickling ≥ LIKELY_OVERTLY_MAL', 'fickling_strict', False),
       ('weights-only all-rejections', 'weights_only_refuse_all', False)]
HEAD = ['Tool', 'Judged T', 'Judged F', 'Inc', 'TP', 'FN', 'FP', 'TN',
        'Recall', 'FNR', 'FPR', 'Precision', 'F1', 'Accuracy']


def confusion(recs, key):
    tp = fn = fp = tn = inc_m = inc_b = 0
    for r in recs:
        gt = r['label']
        if gt not in ('T', 'F'):
            continue
        v = r.get(key)
        na = (v == NA or v is None)
        if gt == 'T':
            if v == 1:
                tp += 1
            else:
                fn += 1
                if na:
                    inc_m += 1
        else:
            if v == 1:
                fp += 1
            elif v == 0:
                tn += 1
            else:
                inc_b += 1
    return tp, fn, fp, tn, inc_m, inc_b


def metrics(tp, fn, fp, tn):
    def d(a, b):
        return round(a / b, 3) if b else None
    return dict(recall=d(tp, tp + fn), fnr=d(fn, tp + fn), fpr=d(fp, fp + tn),
                prec=d(tp, tp + fp), f1=d(2 * tp, 2 * tp + fp + fn),
                acc=d(tp + tn, tp + fn + fp + tn))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--final', required=True)
    ap.add_argument('--out', required=True)
    a = ap.parse_args()
    recs = [json.loads(l) for l in open(a.final, encoding='utf-8') if l.strip()]

    subsets = [('Main table — Real-world (RQ3 comparison target)', [r for r in recs if r['group'] == 'Real-world']),
               ('Auxiliary table 1 — directly generated gadget variants (malicious)', [r for r in recs if r['group'] == 'Gadget variants']),
               ('Auxiliary table 2 — Combined (Real-world + gadget variants)', recs)]

    wb = Workbook()
    HDR = PatternFill('solid', fgColor='DDEBF7'); SUB = PatternFill('solid', fgColor='F2F2F2')
    HL = PatternFill('solid', fgColor='E2EFDA'); GT = PatternFill('solid', fgColor='FFF2CC')
    bd = Border(*[Side('thin', color='BFBFBF')] * 4)
    ws = wb.active; ws.title = 'Results table'
    ws['A1'] = 'RQ3 detection performance — RoPS vs 4 SOTA tools (Main table Real-world 304 · Auxiliary gadget 115 · Combined 419)'
    ws['A1'].font = Font(AR, bold=True, size=13)
    ws['A2'] = ('T=malicious · F=benign. Judged T/F = number of files the tool judged as T/F, Inc = inconclusive (n/a). '
                'See the [Metric definitions] sheet for metric formulas, judgment criteria, and the Inc definition.')
    ws['A2'].font = Font(AR, italic=True, size=9)

    r0 = 4

    def write_block(r0, title, recs_sub):
        nT = sum(1 for r in recs_sub if r['label'] == 'T')
        nF = sum(1 for r in recs_sub if r['label'] == 'F')
        ws.cell(r0, 1, f'{title} — inspected {len(recs_sub)} · GT T(malicious) {nT} · F(benign) {nF}').font = Font(AR, bold=True, size=11)
        h = r0 + 1
        for j, c in enumerate(HEAD, 1):
            x = ws.cell(h, j, c); x.font = Font(AR, bold=True); x.fill = HDR; x.border = bd
            x.alignment = Alignment('center', 'center', wrap_text=True)
        # ground-truth (GT) reference row
        rr = h + 1
        gtrow = ['GT', nT, nF, 0, '', '', '', '', '', '', '', '', '', '']
        for j, v in enumerate(gtrow, 1):
            x = ws.cell(rr, j, v); x.border = bd; x.font = Font(AR, italic=True); x.fill = GT
            if j >= 2:
                x.alignment = Alignment('center')
        rr += 1
        for disp, key, hl in TOOLS + [(None, None, None)] + ALT:
            if disp is None:
                ws.cell(rr, 1, '— Auxiliary (alternate cutoffs) —').font = Font(AR, italic=True, size=9)
                rr += 1
                continue
            tp, fn, fp, tn, inc_m, inc_b = confusion(recs_sub, key)
            m = metrics(tp, fn, fp, tn)
            pt = tp + fp                       # tool judged as T
            pf = tn + (fn - inc_m)             # tool judged as F (incl. GT-malicious seen as F)
            inc = inc_m + inc_b
            inc_str = f'{inc}' + (f' (mal{inc_m}/ben{inc_b})' if inc else '')
            vals = [disp, pt, pf, inc_str, tp, fn, fp, tn,
                    m['recall'], m['fnr'], m['fpr'] if (fp + tn) else '—',
                    m['prec'] if (tp + fp) else '—', m['f1'], m['acc']]
            for j, v in enumerate(vals, 1):
                x = ws.cell(rr, j, v); x.border = bd; x.font = Font(AR, bold=hl)
                if j >= 2:
                    x.alignment = Alignment('center')
                if 9 <= j <= 14 and isinstance(v, float):
                    x.number_format = '0.000'
                if hl:
                    x.fill = HL
            rr += 1
        return rr + 1

    r = r0
    for title, sub in subsets:
        r = write_block(r, title, sub)

    for j, w in enumerate([24, 8, 8, 14, 6, 6, 6, 6, 8, 8, 8, 8, 8, 8], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── Metric definitions sheet ──
    wd = wb.create_sheet('Metric definitions')
    wd['A1'] = 'RQ3 metric definitions · judgment criteria · formulas'
    wd['A1'].font = Font(AR, bold=True, size=13)
    row = 3

    def sec(title):
        nonlocal row
        wd.cell(row, 1, title).font = Font(AR, bold=True, size=11); row += 1

    def line(a_, b_):
        nonlocal row
        x = wd.cell(row, 1, a_); x.font = Font(AR, bold=True); x.alignment = Alignment(vertical='top')
        y = wd.cell(row, 2, b_); y.font = Font(AR); y.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    sec('1. Ground-truth (GT) labels')
    line('T (malicious)', 'Actual malicious file. Positive.')
    line('F (benign)', 'Actual benign file. Negative.')
    line('Total T/F', 'The "GT T(malicious) N · F(benign) M" in each table title is that table\'s total ground-truth count.')
    row += 1

    sec('2. Per-tool judgment criteria (T = counted as malicious detection)')
    line('RoPS (main)', 'warn-or-above = ≥1 of unsafe + review (rops_warn=True) → T')
    line('RoPS unsafe-only (auxiliary)', 'unsafe ≥1 (rops_confirm=True) → T')
    line('picklescan / modelscan', 'tool verdict=malicious → T. But if unreached (status ok · oracle globals>0 · saw 0) then inconclusive (Inc)')
    line('fickling (main/auxiliary)', 'severity idx ≥ SUSPICIOUS(main) / ≥ LIKELY_OVERTLY_MALICIOUS(auxiliary) → T. If status≠ok then Inc')
    line('weights-only (main)', 'loading blocked by Unsupported global (blocked_global) → T, normal load (ok) → F, parse failure → Inc')
    line('weights-only all-rejections (auxiliary)', 'loose cutoff treating everything including parse failures as rejected (T)')
    row += 1

    sec('3. Judged T/F/Inc columns')
    line('Judged T', 'number of files the tool judged as T(malicious) = TP + FP')
    line('Judged F', 'number of files the tool judged as F(benign) = TN + (GT-malicious seen as F)')
    line('Inc (inconclusive)', "number of files where the tool produced no T/F (n/a). 'Inc mal N / ben M' = of those, N GT-malicious · M benign.")
    line('  → handling of mal (N)', 'GT-malicious but inconclusive → the malicious was not caught, so it is counted into FN.')
    line('  → handling of ben (M)', 'GT-benign but inconclusive → since there is no judgment at all, it goes into neither FP nor TN and is excluded from the denominator (prevents a blind tool from getting a free TN).')
    line('Verification identity', 'Judged T + Judged F + Inc = GT T + GT F (all graded items)')
    row += 1

    sec('4. Confusion matrix (Positive=malicious)')
    line('TP', 'GT-malicious ∧ Judged T (malicious correctly called malicious)')
    line('FN', 'GT-malicious ∧ Judged ≠ T (missed a malicious — includes inconclusive)')
    line('FP', 'GT-benign ∧ Judged T (benign falsely called malicious)')
    line('TN', 'GT-benign ∧ Judged F (benign correctly called benign)')
    row += 1

    sec('5. Metric formulas')
    line('Recall (Recall/TPR)', 'TP / (TP + FN) — what % of actual malicious were caught')
    line('FNR (miss rate)', 'FN / (TP + FN) = 1 − Recall — what % of actual malicious were missed')
    line('FPR (false-alarm rate)', 'FP / (FP + TN) — what % of actual benign were called malicious')
    line('Precision', 'TP / (TP + FP) — what % of those called malicious are actually malicious')
    line('F1', '2·TP / (2·TP + FP + FN) — harmonic mean of Precision and Recall')
    line('Accuracy', '(TP + TN) / (TP + FN + FP + TN) — fraction correct over all')
    line('Denominator note', 'FPR·Precision are — (undefined) when the denominator is 0. Gadget variants (all malicious) have benign=0, so no FPR·TN.')
    wd.column_dimensions['A'].width = 26
    wd.column_dimensions['B'].width = 88
    for rr in range(1, row):
        wd.cell(rr, 2).alignment = Alignment(wrap_text=True, vertical='top')

    # ── fickling(blob) Real-world dedicated sheet ──
    wbl = wb.create_sheet('fickling(blob) real-world')
    rw = [r for r in recs if r['group'] == 'Real-world']
    have = [r for r in rw if r.get('has_blob_data')]
    miss = [r['sample_id'] for r in rw if not r.get('has_blob_data')]
    nT = sum(1 for r in rw if r['label'] == 'T'); nF = sum(1 for r in rw if r['label'] == 'F')
    wbl['A1'] = 'RQ3 auxiliary — fickling(blob) × RoPS · Real-world 304'
    wbl['A1'].font = Font(AR, bold=True, size=13)
    npkl = [r['sample_id'] for r in rw if r.get('fickling_blob_status') == 'no_pickle']
    wbl['A2'] = ('Result of running fickling on each blob (pickle stream) that RoPS carved. The constraint of '
                 'not being able to open containers/streams from file input is removed here (= RoPS+fickling combination). '
                 f'Measured {len(have)}/{len(rw)}. pickle absent {len(npkl)} (nothing to scan → Inc, justified)'
                 + (f' · unmeasured {len(miss)} (original read failure, footnote below).' if miss else ' · unmeasured 0 (all obtained).'))
    wbl['A2'].font = Font(AR, italic=True, size=9)
    wbl['A3'] = ('T=malicious. severity ≥ SUSPICIOUS(main) / ≥ LIKELY_OVERTLY_MALICIOUS(auxiliary) → T. '
                 'If there is no blob (pickle absent) or it is unmeasured, Inc. Metric definitions in the [Metric definitions] sheet.')
    wbl['A3'].font = Font(AR, italic=True, size=9)
    hh = 5
    for j, c in enumerate(HEAD, 1):
        x = wbl.cell(hh, j, c); x.font = Font(AR, bold=True); x.fill = HDR; x.border = bd
        x.alignment = Alignment('center', 'center', wrap_text=True)
    # ground-truth row
    for j, v in enumerate(['GT', nT, nF, 0] + [''] * 10, 1):
        x = wbl.cell(hh + 1, j, v); x.border = bd; x.font = Font(AR, italic=True); x.fill = GT
        if j >= 2:
            x.alignment = Alignment('center')
    blrows = [('fickling(blob) ≥ SUSPICIOUS', 'fickling_blob_loose', True),
              ('fickling(blob) ≥ LIKELY_OVERTLY_MAL', 'fickling_blob_strict', False),
              ('(reference) fickling(file) ≥ SUSPICIOUS', 'fickling_loose', False)]
    rr = hh + 2
    for disp, key, hl in blrows:
        tp, fn, fp, tn, inc_m, inc_b = confusion(rw, key)
        m = metrics(tp, fn, fp, tn)
        pt = tp + fp; pf = tn + (fn - inc_m); inc = inc_m + inc_b
        inc_str = f'{inc}' + (f' (mal{inc_m}/ben{inc_b})' if inc else '')
        vals = [disp, pt, pf, inc_str, tp, fn, fp, tn, m['recall'], m['fnr'],
                m['fpr'] if (fp + tn) else '—', m['prec'] if (tp + fp) else '—', m['f1'], m['acc']]
        for j, v in enumerate(vals, 1):
            x = wbl.cell(rr, j, v); x.border = bd; x.font = Font(AR, bold=hl)
            if j >= 2:
                x.alignment = Alignment('center')
            if 9 <= j <= 14 and isinstance(v, float):
                x.number_format = '0.000'
            if hl:
                x.fill = HL
        rr += 1
    rr += 1
    wbl.cell(rr, 1, f'pickle absent {len(npkl)} (no blob to scan → Inc, justified): ' + ' · '.join(npkl)
             + ' — pure tensor format (0 pickle protos on decompress, original eval or_globals=0). 3-3-D-00010 (malicious) has no pickle due to tar path traversal.'
             ).font = Font(AR, italic=True, size=9)
    rr += 1
    if miss:
        wbl.cell(rr, 1, f'blob unmeasured {len(miss)} (benign, treated as Inc): ' + ' · '.join(miss)
                 + ' — the 1.7GB .bin originals failed to read from the mount, not obtained this session. '
                   'Can be filled by extracting zip data.pkl on the device then running fickling (README). '
                   'Since the original-blob benign 214/215 are flagged, even if filled it is likely FP (FPR≈1.0).'
                 ).font = Font(AR, italic=True, size=9)
        rr += 1
    wbl.cell(rr, 1, '※ fickling(blob) flags almost all large benign-model pickles as resource-limit / expansion-attack class → extreme FPR. '
             'Shows that RoPS carving raises fickling\'s reachability (Recall↑) but does not improve its benign-discrimination ability.'
             ).font = Font(AR, italic=True, size=9)
    for j, w in enumerate([30, 8, 8, 14, 6, 6, 6, 6, 8, 8, 8, 8, 8, 8], 1):
        wbl.column_dimensions[get_column_letter(j)].width = w

    # ── Analysis ──
    ws2 = wb.create_sheet('Analysis')
    cols = ['sample_id', 'tier', 'GT', 'group', 'origin', 'RoPS', 'picklescan', 'modelscan',
            'fickling_loose', 'fickling_blob_loose', 'has_blob_data', 'weights_only',
            'RoPS_confirm', 'fickling_strict', 'wo_status', 'fk_severity', 'or_globals']
    for j, c in enumerate(cols, 1):
        x = ws2.cell(1, j, c); x.font = Font(AR, bold=True); x.fill = HDR; x.border = bd
        x.alignment = Alignment('center', 'center')
    # Display-only English labels for the internal `group` codes.
    GROUP_DISPLAY = {'Real-world': 'Real-world', 'Gadget variants': 'Gadget variants'}
    for i, r_ in enumerate(recs, 2):
        for j, c in enumerate(cols, 1):
            v = r_.get(c, '')
            if c == 'group':
                v = GROUP_DISPLAY.get(v, v)
            ws2.cell(i, j, v).font = Font(AR, size=9)
    for j, w in enumerate([15, 7, 5, 10, 14, 7, 10, 10, 12, 14, 12, 12, 12, 12, 16, 22, 10], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = 'B2'
    wb.save(a.out)

    # console summary
    for title, sub in subsets:
        nT = sum(1 for r in sub if r['label'] == 'T'); nF = sum(1 for r in sub if r['label'] == 'F')
        print(f'\n[{title}] n={len(sub)} GT T{nT}/F{nF}')
        for disp, key, _ in TOOLS:
            tp, fn, fp, tn, im, ib = confusion(sub, key)
            m = metrics(tp, fn, fp, tn)
            print(f'  {disp:24} JudgedT{tp+fp} JudgedF{tn+(fn-im)} Inc{im+ib}(mal{im}/ben{ib}) | '
                  f'TP{tp} FN{fn} FP{fp} TN{tn} recall={m["recall"]} FPR={m["fpr"]}')
    print('\n->', a.out)


if __name__ == '__main__':
    main()
