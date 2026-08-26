#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RQ2 (denylist V2) final integrated jsonl generation — self-contained.

RQ2 uses 'public denylist (picklescan∪modelscan) ∩ actual benign-model use' as its axis,
showing that RoPS reduces list-based false positives by **moving the judgment basis to the argument**.

Axis (denylisted item) attribution is determined by the **harness trace's globals (all referenced names)**.
This is only reproducible on the user's Mac (~/rops/harness/traces), so two paths are provided.

  --harness <dir>    : compute axes directly from harness traces (original pipeline, full reproduction)
  --attrib <xlsx>    : reproduce by reading the resolved axis attribution from the 'File_Attribution' sheet
                       of the existing authoritative table (RQ2_authoritative.xlsx) (same result without traces)

Both paths merge with rq2_results.jsonl (judgments) and MASTER_INDEX.csv (metadata) to
build rq2_final.jsonl. When the corpus is re-adjusted, the table is re-aggregated from this file alone.
"""
import argparse, json, os, csv, collections

BEN = {'2-A', '2-B', '2-C', '2-D'}
MAL = {'3-A', '3-B', '3-C', '3-D', '3-E', '3-G'}


def in_main(r):
    """RQ2 redesigned corpus: status ok, scope in, exclude 3-E (dev catalog)·1-C (probe), include 2-C (dev)."""
    if r.get('status', 'ok') != 'ok':
        return False
    if r['tier'] in ('3-E', '1-C'):
        return False
    return (r.get('scope') or 'in') == 'in'


def load_deny(p):
    d = json.load(open(p, encoding='utf-8'))
    entries = set(d['picklescan']) | set(d['modelscan'])
    exact = {x for x in entries if not x.endswith('.*')}
    pref = sorted({x[:-2] for x in entries if x.endswith('.*')}, key=len, reverse=True)
    return exact, pref


def make_norm():
    try:
        import _compat_pickle as _CP
        IM, NM = dict(_CP.IMPORT_MAPPING), dict(_CP.NAME_MAPPING)
    except Exception:
        IM, NM = {}, {}

    def norm(x):
        if not x:
            return ''
        x = x.strip().replace(' ', '.')
        mod, _, name = x.rpartition('.')
        if not mod:
            return x
        if (mod, name) in NM:
            mod, name = NM[(mod, name)]
        elif mod in IM:
            mod = IM[mod]
        return f'{mod}.{name}'
    return norm


def labeler(exact, pref):
    def f(c):
        if c in exact:
            return c
        m = c.rsplit('.', 1)[0]
        for p in pref:
            if m == p or m.startswith(p + '.'):
                return p + '.*'
        return None
    return f


def tags_from_harness(harness, sid, lab, norm):
    p = os.path.join(harness, 'traces', sid + '.json')
    if not os.path.exists(p):
        return None
    try:
        t = json.load(open(p, encoding='utf-8'))
    except Exception:
        return None
    names = set()
    for s in t.get('streams', []):
        for g in s.get('globals') or []:
            names.add(norm('%s.%s' % (g.get('module'), g.get('name'))))
    tags = {lab(c) for c in names}
    tags.discard(None)
    return tags


def attrib_from_xlsx(path):
    """'File_Attribution' sheet of the existing authoritative table -> {sid: denylisted item}. (resolved axis attribution)"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb['File_Attribution']  # sheet name in the authoritative source workbook
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None:
            continue
        entry, side, sid = row[0], row[1], row[2]
        out[sid] = entry
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--res', required=True, help='rq2_results.jsonl (V2)')
    ap.add_argument('--ledger', required=True)
    ap.add_argument('--deny', required=True, help='sota_denylist.json')
    ap.add_argument('--harness', default='', help='harness trace directory (full reproduction)')
    ap.add_argument('--attrib', default='', help='existing RQ2_authoritative.xlsx (file-attribution reproduction)')
    ap.add_argument('--out', required=True)
    a = ap.parse_args()

    led = {r['sample_id']: r for r in csv.DictReader(open(a.ledger, encoding='utf-8'))}
    R = {json.loads(l)['sample_id']: json.loads(l)
         for l in open(a.res, encoding='utf-8') if l.strip()}
    lab = labeler(*load_deny(a.deny))
    norm = make_norm()

    xattr = attrib_from_xlsx(a.attrib) if a.attrib else {}
    if not a.harness and not a.attrib:
        raise SystemExit('One of --harness or --attrib is required.')

    # pass 1: collect the set of denylisted items (tags) per file
    tagmap = {}
    order = []
    for sid in sorted(R):
        r = R[sid]
        if not in_main(r):
            continue
        order.append(sid)
        if a.harness:
            tset = tags_from_harness(a.harness, sid, lab, norm)
            tagmap[sid] = sorted(tset) if tset else []
        else:
            e = xattr.get(sid)
            tagmap[sid] = [e] if e else []
    # item frequency (all) -> single attribution to the rarest item (same as original pipeline)
    freq = collections.Counter()
    for sid in order:
        freq.update(tagmap[sid])

    out = open(a.out, 'w', encoding='utf-8')
    n = 0
    for sid in order:
        r = R[sid]
        L = led.get(sid, {})
        tier = r['tier']
        side = 'benign' if tier in BEN else ('malicious' if tier in MAL else 'other')
        tags = tagmap[sid]
        entry = min(tags, key=lambda c: (freq[c], c)) if tags else None   # canonical attribution

        rec = {
            'sample_id': sid,
            'tier': tier,
            'origin': L.get('origin', ''),
            'label': L.get('label', r.get('label', '')),
            'format': r.get('format', L.get('format', '')),
            'side': side,
            'metric_role': r.get('metric_role', ''),
            'level': r.get('level', ''),
            'variant_kind': r.get('variant_kind', ''),
            'gadget_id': r.get('gadget_id', ''),
            'canary_fired': r.get('canary_fired', ''),
            # axis attribution
            'denylist_entry': entry,           # denylisted item the file is attributed to (null if none)
            'denylist_tags': tags,             # (harness mode) full set of denylisted items used
            # RoPS judgment (raw data for re-aggregation)
            'H1': bool(r.get('H1')),           # Review (unsafe+review >=1)
            'H1_cf': bool(r.get('H1_cf')),
            'n_unsafe': int(r.get('n_unsafe') or 0),   # Unsafe (unsafe)
            'n_review': int(r.get('n_review') or 0),
            'n_low': int(r.get('n_low') or 0),
            'n_hits': int(r.get('n_hits') or 0),
            'callables': r.get('callables') or [],
        }
        out.write(json.dumps(rec, ensure_ascii=False) + '\n')
        n += 1
    out.close()

    recs = [json.loads(l) for l in open(a.out, encoding='utf-8') if l.strip()]
    dl = [r for r in recs if r['denylist_entry']]
    cb = collections.Counter(r['side'] for r in dl)
    print(f'rq2_final.jsonl: {n} in_main records -> {a.out}')
    print(f'  denylist-attributed files {len(dl)}  (benign {cb.get("benign",0)} · malicious {cb.get("malicious",0)})')
    print(f'  denylisted item count {len(set(r["denylist_entry"] for r in dl))}')


if __name__ == '__main__':
    main()
